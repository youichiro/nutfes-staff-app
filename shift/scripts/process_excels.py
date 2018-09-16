"""Memo: 同じディレクトリにシフトのxlsxファイルを配置する"""
import os
import openpyxl
from tqdm import tqdm
from shift.scripts.const import row_to_time_str
from shift.models import Shift


def get_value_list(tuple_2d):
    return [[cell.value for cell in row] for row in tuple_2d]


def save(user, department, day, weather, shift_id, row, task):
    user = user.replace(' ', '')
    task = task.replace(' ', '').replace('\n', '') if task else None
    time_str = row_to_time_str[row]
    shift, is_created = Shift.objects.get_or_create(user=user, department=department,
                                                    day=day, weather=weather, shift_id=shift_id)
    if shift.__getattribute__(time_str):
        return
    shift.__setattr__(time_str, task)
    shift.save()


def register(sheet, end_column, day, weather, shift_id, active=False):
    if not active:
       return
    print('Saving: {}{}_shift'.format(day, weather))

    # 列と名前の対応表の作成
    col_to_name = {}
    name_cells = sheet['B3:'+end_column+'3'][0]
    for cell in name_cells:
        name = cell.value
        col = cell.col_idx
        col_to_name[col] = name

    # 列と局の対応表の作成
    col_to_depart = {}
    depart_cells = sheet['B2:'+end_column+'2'][0]
    depart = depart_cells[0].value
    for cell in depart_cells:
        depart = cell.value or depart
        col = cell.col_idx
        col_to_depart[col] = depart

    # 結合セルのシフト登録
    merged_cells = sheet.merged_cells.ranges
    for merged_cell in tqdm(merged_cells):
        cell_range = merged_cell.ref
        cells = sheet[cell_range]
        task = get_value_list(cells)[0][0]
        for cell in cells:
            cell = cell[0]
            row = cell.row
            col = cell.col_idx

            # シフト範囲外は除く
            if row not in row_to_time_str.keys():
                continue
            if col not in col_to_name.keys():
                continue

            user = col_to_name[col]
            department = col_to_depart[col]
            save(user, department, day, weather, shift_id, row, task)

    # 結合セル以外のシフト登録
    all_cells = sheet['B4:'+end_column+'38']
    for row_cells in tqdm(all_cells):
        for cell in row_cells:
            row = cell.row
            col = cell.col_idx
            user = col_to_name[col]
            department = col_to_depart[col]
            task = cell.value
            save(user, department, day, weather, shift_id, row, task)


def main():
    # files
    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    first_shift_file = os.path.join(current_dir, 'scripts/first_shift.xlsx')
    second_shift_file = os.path.join(current_dir, 'scripts/second_shift.xlsx')
    preparation_shift_file = os.path.join(current_dir, 'scripts/preparation_shift.xlsx')
    # cleanup_shift_file = os.path.join(current_dir, 'scripts/cleanup_shift.xlsx')

    # Workbooks
    first_wb = openpyxl.load_workbook(first_shift_file)
    second_wb = openpyxl.load_workbook(second_shift_file)
    preparation_wb = openpyxl.load_workbook(preparation_shift_file)
    # cleanup_wb = openpyxl.load_workbook(cleanup_shift_file)

    # Worksheets
    first_sun_sheet = first_wb['晴 ver.2.0']
    first_rain_sheet = first_wb['雨 ver.2.0']
    second_sun_sheet = second_wb['晴 ver.2.0']
    second_rain_sheet = second_wb['雨 ver.2.0']
    # preparation_sun_sheet = preparation_wb[preparation_wb.sheetnames[0]]
    preparation_sun_sheet = preparation_wb['準備日 晴 当日変更']
    # preparation_rain_sheet = preparation_wb[preparation_wb.sheetnames[1]]
    # cleanup_sheet = cleanup_wb[cleanup_wb.sheetnames[0]]

    # Delete all
    Shift.objects.all().delete()

    # Registration
    register(preparation_sun_sheet, end_column='DJ', day='準備日', weather='晴', shift_id=1, active=True)
    # register(preparation_rain_sheet, end_column='DJ', day='準備日', weather='雨', shift_id=2, active=False)
    register(first_sun_sheet, end_column='DR', day='1日目', weather='晴', shift_id=3, active=True)
    register(first_rain_sheet, end_column='DS', day='1日目', weather='雨', shift_id=4, active=True)
    register(second_sun_sheet, end_column='DE', day='2日目', weather='晴', shift_id=5, active=True)
    register(second_rain_sheet, end_column='DE', day='2日目', weather='雨', shift_id=6, active=True)
    # register(cleanup_sheet, end_column='EQ', day='片付け日', weather='', shift_id=7, active=False)

    print('Success saving all.')
